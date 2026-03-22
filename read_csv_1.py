# -*- coding: utf-8 -*-
from collections import defaultdict
import csv
import os
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment

# Directory containing the CSV bank statements and for the Excel output
DATA_DIR = r"D:\Documents\formalités\compte perso"

def get_category(label_parts):
    """
    Extracts a category from label parts.
    Strategy: 
    - If there are 4+ parts (Type - Label - Code - Cat), the last one is the category.
    - If there are 3 parts (Type - Label - Cat), the last one is the category.
    - Otherwise, use the Type (first part) as category.
    """
    if not label_parts:
        return "Divers"
    if len(label_parts) >= 3:
        return label_parts[-1]
    return label_parts[0]

def export_to_excel(transactions, monthly_summary, output_file):
    wb = Workbook()
    
    # --- Onglet Synthèse ---
    ws_summary = wb.active
    ws_summary.title = "Synthèse"
    
    headers = ["Mois", "Evolution Totale", "Solde Final"]
    ws_summary.append(headers)
    
    # Style headers
    for cell in ws_summary[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")

    for row in monthly_summary:
        ws_summary.append(row)
    
    # Format colonnes
    for row in ws_summary.iter_rows(min_row=2, max_col=3):
        row[1].number_format = '#,##0.00'
        row[2].number_format = '#,##0.00'

    # --- Onglet Statistiques ---
    ws_stats = wb.create_sheet(title="Statistiques")
    
    # 1. Préparer les données
    all_months = sorted(set(tx['Month'] for tx in transactions))
    all_categories = sorted(set(tx.get('Categorie', 'Divers') for tx in transactions))
    
    # Identifier les mois complets pour les moyennes
    full_months = [m for m in all_months if any(other > m for other in all_months)]
    months_covered = len(full_months)
    
    if months_covered == 0:
        months_covered = 1
        stats_note = f"Moyennes calculées sur le seul mois disponible ({all_months[0]})"
    else:
        stats_note = f"Moyennes mensuelles calculées sur {months_covered} mois complets : {', '.join(full_months)}"

    # 2. En-têtes (Catégorie | Mois1 | Mois2 | ... | Moyenne)
    headers = ["Poste de dépense"] + all_months + [f"Moyenne ({months_covered} mois)"]
    ws_stats.append(headers)
    
    # Style en-têtes
    for cell in ws_stats[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")

    # 3. Calculs des totaux par (catégorie, mois)
    # Données : cat_month_data[category][month] = total_amount
    cat_month_data = defaultdict(lambda: defaultdict(float))
    for tx in transactions:
        cat = tx.get('Categorie', 'Divers')
        month = tx['Month']
        cat_month_data[cat][month] += tx['Montant']

    # 4. Remplir les lignes
    for cat in all_categories:
        row = [cat]
        sum_full_months = 0.0
        
        for month in all_months:
            amount = cat_month_data[cat][month]
            row.append(amount)
            if month in full_months:
                sum_full_months += amount
        
        # Ajouter la moyenne
        row.append(sum_full_months / months_covered)
        ws_stats.append(row)

    # 5. Note explicative en bas
    ws_stats.append([])
    ws_stats.append([stats_note])

    # 6. Formatage
    # Nombre de colonnes : 1 (cat) + len(all_months) + 1 (moyenne)
    num_cols = len(all_months) + 2
    # Adjust format for numbers only (skip cat name in col A)
    for row in ws_stats.iter_rows(min_row=2, max_row=len(all_categories) + 1, min_col=2, max_col=num_cols):
        for cell in row:
            cell.number_format = '#,##0.00'
            
    ws_stats.column_dimensions['A'].width = 30
    for i in range(2, num_cols + 1):
        col_letter = ws_stats.cell(row=1, column=i).column_letter
        ws_stats.column_dimensions[col_letter].width = 15

    # --- Onglets Mensuels ---
    # Group transactions for easier sheet creation
    by_month = defaultdict(list)
    for tx in transactions:
        by_month[tx['Month']].append(tx)
        
    for month in sorted(by_month.keys(), reverse=True):
        ws = wb.create_sheet(title=month)
        ws.append(["Date", "Catégorie", "Libellé", "Montant"])
        
        # Style headers
        for cell in ws[1]:
            cell.font = Font(bold=True)
            
        for tx in by_month[month]:
            ws.append([tx['Date'], tx.get('Categorie', ''), tx['Libelle'], tx['Montant']])
            
        # Format amount column
        for row in ws.iter_rows(min_row=2, max_col=4):
            row[3].number_format = '#,##0.00'
            
        # Ajuster les largeurs de colonnes
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 50
        ws.column_dimensions['D'].width = 15

    try:
        wb.save(output_file)
        print(f"\nFichier Excel créé : {output_file}")
    except PermissionError:
        print(f"\nERREUR : Impossible d'enregistrer le fichier '{output_file}'.")
        print("Veuillez vous assurer que le fichier n'est pas ouvert dans Excel, puis relancez le script.")

def merge_csv_files():
    """
    Merges all CSV files in DATA_DIR into fusion.csv.
    Uses the newest file's data for any given date.
    Deduplicates identical transactions.
    Adapted for 'compte perso' format (no header, specific columns).
    """
    if not os.path.exists(DATA_DIR):
        print(f"Error: Directory not found: {DATA_DIR}")
        return None

    csv_files = [f for f in os.listdir(DATA_DIR) if f.lower().endswith('.csv') and f.lower() != 'fusion.csv']
    
    if not csv_files:
        print("Aucun fichier CSV à fusionner.")
        return None

    # Sort files by modification time (oldest first)
    file_info = []
    for f in csv_files:
        path = os.path.join(DATA_DIR, f)
        file_info.append({
            'name': f,
            'path': path,
            'mtime': os.path.getmtime(path)
        })
    
    file_info.sort(key=lambda x: x['mtime'])
    
    date_registry = {}
    latest_balance = 0.0
    latest_mtime = -1

    for info in file_info:
        with open(info['path'], mode='r', encoding='utf-8-sig') as f:
            lines = f.readlines()
            if not lines:
                continue
            
            # Identify balance from the LAST line of the most recent file
            # Format in 'compte perso': Date;Amount;;Compte
            last_line = lines[-1].strip().split(';')
            if info['mtime'] > latest_mtime:
                latest_mtime = info['mtime']
                if len(last_line) >= 2:
                    try:
                        latest_balance = float(last_line[1].replace(',', '.'))
                    except ValueError:
                        pass
            
            # Read transactions
            for line in lines:
                parts = line.strip().split(';')
                # Transaction lines have more columns (~8) than balance lines (4)
                if len(parts) < 5:
                    continue
                
                date = parts[0]
                amount_str = parts[1] # Garder tel quel pour fusion.csv
                
                # Récupérer toutes les parties du libellé (type + descriptions diverses)
                # On commence à l'index 2 (Type) et on prend tout ce qui n'est pas vide
                label_parts = []
                for i in range(2, len(parts)):
                    p = parts[i].strip()
                    if p and p not in label_parts:
                        label_parts.append(p)
                
                full_label = " - ".join(label_parts)
                tx_tuple = (full_label, amount_str)
                
                if date not in date_registry or info['mtime'] > date_registry[date]['mtime']:
                    date_registry[date] = {
                        'mtime': info['mtime'],
                        'transactions': {tx_tuple}
                    }
                elif info['mtime'] == date_registry[date]['mtime']:
                    date_registry[date]['transactions'].add(tx_tuple)

    # Sort dates
    def parse_date(d):
        try:
            parts = d.split('/')
            return f"{parts[2]}-{parts[1]}-{parts[0]}"
        except:
            return d

    sorted_dates = sorted(date_registry.keys(), key=parse_date)
    
    fusion_path = os.path.join(DATA_DIR, "fusion.csv")
    print(f"Génération de fusion.csv dans {DATA_DIR}...")
    
    try:
        with open(fusion_path, mode='w', encoding='utf-8-sig', newline='') as f:
            # Reconstruct a header compatible with read_bank_csv
            f.write(";;;\n;;;\n;;;\n;;;\n")
            # Line 5: Balance
            f.write(f"Solde (EUROS);;;{str(latest_balance).replace('.', ',')}\n")
            # Line 6: Empty
            f.write(";;;\n")
            # Line 7: Header
            f.write("Date;Libellé;Montant(EUROS)\n")
            
            # Write transactions
            for date in sorted_dates:
                for tx_tuple in sorted(date_registry[date]['transactions']):
                    f.write(f"{date};{tx_tuple[0]};{tx_tuple[1]}\n")
    except PermissionError:
        print(f"\nERREUR : Impossible d'écrire dans '{fusion_path}'.")
        print("Veuillez vous assurer que le fichier n'est pas ouvert dans Excel ou un autre programme.")
        return None
    
    return fusion_path

def read_bank_csv(file_path):
    if not file_path or not os.path.exists(file_path):
        print(f"CSV file not found: {file_path}")
        return None

    current_balance = 0.0

    with open(file_path, mode='r', encoding='utf-8-sig') as f:
        # Read the first few lines to extract balance (line 5)
        lines = f.readlines()
        if len(lines) < 5:
            print("Invalid CSV format (too short).")
            return
            
        balance_line = lines[4].strip()
        if balance_line.startswith("Solde"):
            balance_str = balance_line.split(';')[-1].replace(',', '.')
            try:
                current_balance = float(balance_str)
            except ValueError:
                current_balance = 0.0
            print(f"Solde initial (depuis l'en-tête) : {current_balance:.2f} EUROS")
            print("-" * 30)

    # Dictionary to store monthly totals
    monthly_totals = defaultdict(float)

    # Ré-ouverture pour DictReader
    with open(file_path, mode='r', encoding='utf-8-sig') as f:
        # Skip metadata header (first 6 lines)
        for _ in range(6):
            next(f)
            
        reader = csv.DictReader(f, delimiter=';')
        
        transactions = []
        for row in reader:
            if not row.get('Date'):
                continue
                
            date_str = row.get('Date')
            
            try:
                parts = date_str.split('/')
                month_key = f"{parts[2]}-{parts[1]}"
            except (IndexError, AttributeError):
                continue

            try:
                raw_amount = row['Montant(EUROS)'].replace(',', '.')
                amount = float(raw_amount)
            except (ValueError, KeyError, AttributeError):
                amount = 0.0
                
            transaction = {
                'Date': date_str,
                'Month': month_key,
                'Libelle': row.get('Libellé'),
                'Categorie': get_category(row.get('Libellé', '').split(' - ')),
                'Montant': amount
            }
            transactions.append(transaction)
            monthly_totals[month_key] += amount

    # Afficher la synthèse par mois
    print("\nSynthèse Mensuelle :")
    print(f"{'Mois':<10} | {'Evolution':>15} | {'Solde Final':>15}")
    print("-" * 46)
    
    current_running_balance = current_balance
    sorted_months = sorted(monthly_totals.keys(), reverse=True)
    summary_rows = []
    
    for month in sorted_months:
        total_change = monthly_totals[month]
        balance_at_end = current_running_balance
        summary_rows.append((month, total_change, balance_at_end))
        current_running_balance -= total_change

    for month, total_change, balance_at_end in summary_rows:
        print(f"{month:<10} | {total_change:>15.2f} | {balance_at_end:>15.2f}")

    # Export to Excel
    output_excel = os.path.join(DATA_DIR, "compte_perso.xlsx")
    export_to_excel(transactions, summary_rows, output_excel)

    return transactions, monthly_totals

if __name__ == "__main__":
    fusion_file = merge_csv_files()
    if fusion_file:
        print(f"Fusion terminée. Traitement de {fusion_file}...")
        read_bank_csv(fusion_file)
    else:
        print("Erreur : Impossible de créer le fichier de fusion.")
